"""
NSE Stock Analyzer v6 — Live Data + Top 10 Dashboard
==================================================
• Live data via yfinance ONLY (no simulation)
• 3 Timeframes: 1H, 4H, 1D
• Heikin Ashi candles for all timeframes
• Custom Signal: BUY (RSI>40 & Stoch>20) | SELL (RSI<70 & Stoch<80)
• Top 10 stocks by performance
• Auto-refresh timestamp on every run
"""

import os, sys, warnings
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

warnings.filterwarnings("ignore")

# ── CONFIG ─────────────────────────────────────────────────────────
TIMEFRAMES    = ["1h", "4h", "1d"]
LOOKBACK_DAYS = 60  # Enough for daily data

# NEW SIGNAL RULES (as per your requirement)
BUY_RSI       = 40      # RSI > 40 for BUY
BUY_STOCH     = 20      # Stoch > 20 for BUY
SELL_RSI      = 70      # RSI < 70 for SELL
SELL_STOCH    = 80      # Stoch < 80 for SELL

RSI_PERIOD    = 14
STOCH_K       = 14
STOCH_D       = 3
MACD_FAST     = 12
MACD_SLOW     = 26
MACD_SIG      = 9
BB_PERIOD     = 20
BB_STD        = 2
OUTPUT_FILE   = "NSE_Live_Top10_Signals_v6.xlsx"

NSE_STOCKS = {
    "RELIANCE":   "RELIANCE.NS",
    "TCS":        "TCS.NS",
    "INFY":       "INFY.NS",
    "HDFCBANK":   "HDFCBANK.NS",
    "ICICIBANK":  "ICICIBANK.NS",
    "WIPRO":      "WIPRO.NS",
    "AXISBANK":   "AXISBANK.NS",
    "BAJFINANCE": "BAJFINANCE.NS",
    "MARUTI":     "MARUTI.NS",
    "SUNPHARMA":  "SUNPHARMA.NS",
    "TATAMOTORS": "TATAMOTORS.NS",
    "LTIM":       "LTIM.NS",
    "HCLTECH":    "HCLTECH.NS",
    "KOTAKBANK":  "KOTAKBANK.NS",
    "SBIN":       "SBIN.NS",
}

# ── STYLING CONSTANTS ──────────────────────────────────────────────
HDR_DARK    = "1F3864"
HDR_MED     = "2E4D8A"
BUY_BG      = "C6EFCE"; BUY_FG      = "276221"
SELL_BG     = "FFC7CE"; SELL_FG     = "9C0006"
HOLD_BG     = "FFEB9C"; HOLD_FG     = "9C5700"
WAIT_BG     = "D9D9D9"; WAIT_FG     = "666666"
STR_HI_BG   = "C6EFCE"
STR_MED_BG  = "FFEB9C"
STR_LO_BG   = "FFC7CE"
ALT_BG      = "EEF2FF"
WHITE       = "FFFFFF"

thin = Side(style="thin", color="CCCCCC")
brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

# ── HEIKIN ASHI ────────────────────────────────────────────────────
def to_heikin_ashi(df):
    ha = df.copy()
    ha["HA_Close"] = (df["Open"] + df["High"] + df["Low"] + df["Close"]) / 4
    ha["HA_Open"]  = ha["HA_Close"].copy()
    ha.iloc[0, ha.columns.get_loc("HA_Open")] = (df["Open"].iloc[0] + df["Close"].iloc[0]) / 2
    for i in range(1, len(ha)):
        ha.iloc[i, ha.columns.get_loc("HA_Open")] = (
            ha.iloc[i-1]["HA_Open"] + ha.iloc[i-1]["HA_Close"]) / 2
    ha["HA_High"] = ha[["HA_Open","HA_Close","High"]].max(axis=1)
    ha["HA_Low"]  = ha[["HA_Open","HA_Close","Low"]].min(axis=1)
    return ha

# ── INDICATORS ─────────────────────────────────────────────────────
def calc_rsi(close, period=14):
    delta = close.diff()
    gain  = delta.clip(lower=0).ewm(com=period-1, adjust=False).mean()
    loss  = (-delta.clip(upper=0)).ewm(com=period-1, adjust=False).mean()
    rs    = gain / loss.replace(0, np.nan)
    return 100 - (100 / (1+rs))

def calc_stoch(high, low, close, k=14, d=3):
    ll = low.rolling(k).min()
    hh = high.rolling(k).max()
    pk = 100*(close-ll)/(hh-ll).replace(0, np.nan)
    return pk, pk.rolling(d).mean()

def calc_macd(close, fast=12, slow=26, sig=9):
    ef = close.ewm(span=fast, adjust=False).mean()
    es = close.ewm(span=slow, adjust=False).mean()
    ml = ef - es
    sl = ml.ewm(span=sig, adjust=False).mean()
    return ml, sl, ml - sl

def calc_bb(close, p=20, s=2):
    sma = close.rolling(p).mean()
    std = close.rolling(p).std()
    return sma, sma+s*std, sma-s*std

def calc_ema(close, p=21):
    return close.ewm(span=p, adjust=False).mean()

# NEW SIGNAL FUNCTION - UPDATED AS PER YOUR REQUIREMENTS
def get_signal(rsi, stk):
    if pd.isna(rsi) or pd.isna(stk):
        return "⏳ WAIT"
    # BUY: RSI > 40 AND Stoch > 20
    if rsi > BUY_RSI and stk > BUY_STOCH:
        return "🟢 BUY"
    # SELL: RSI < 70 AND Stoch < 80
    if rsi < SELL_RSI and stk < SELL_STOCH:
        return "🔴 SELL"
    return "🟡 HOLD"

def signal_strength(rsi, stk, macd_h, vol, avg_vol):
    score = 0
    if pd.isna(rsi) or pd.isna(stk):
        return 0
    # BUY strength: higher RSI and Stoch = stronger BUY
    if rsi > 40:
        score += min(35, int((rsi - 40) / 30 * 70))
    # SELL strength: lower RSI and Stoch = stronger SELL
    if rsi < 70:
        score += min(35, int((70 - rsi) / 30 * 70))
    if stk > 20:
        score += min(35, int((stk - 20) / 60 * 70))
    if stk < 80:
        score += min(35, int((80 - stk) / 60 * 70))
    if not pd.isna(macd_h):
        score += 20 if (macd_h > 0 and rsi > 40) or (macd_h < 0 and rsi < 70) else 5
    if not pd.isna(avg_vol) and avg_vol > 0 and vol/avg_vol > 1.5:
        score += 10
    return min(100, score)

def sig_style(sig):
    """Convert signal string to background and foreground colors"""
    if not isinstance(sig, str):
        return WAIT_BG, WAIT_FG
    if "BUY" in sig:
        return BUY_BG, BUY_FG
    if "SELL" in sig:
        return SELL_BG, SELL_FG
    if "HOLD" in sig:
        return HOLD_BG, HOLD_FG
    return WAIT_BG, WAIT_FG

def sig_short(sig):
    if not isinstance(sig, str):
        return "WAIT"
    if "BUY" in sig:
        return "BUY"
    if "SELL" in sig:
        return "SELL"
    if "HOLD" in sig:
        return "HOLD"
    return "WAIT"

# ── FETCH LIVE DATA ONLY (NO SIMULATION) ──────────────────────────
def fetch_live_data(name, ticker, interval):
    """Fetch live data from yfinance - NO SIMULATION"""
    try:
        import yfinance as yf
        df = yf.download(ticker, period=f"{LOOKBACK_DAYS}d",
                         interval=interval, progress=False, auto_adjust=True)
        if df is not None and len(df) > RSI_PERIOD + 5:
            df.columns = [c[0] if isinstance(c,tuple) else c for c in df.columns]
            df = df[~df.index.duplicated(keep="last")].sort_index()
            print(f"  ✅ {name} ({interval}): {len(df)} candles [LIVE]")
            return df, True
        else:
            print(f"  ❌ {name} ({interval}): Insufficient data ({len(df) if df is not None else 0} candles)")
            return None, False
    except Exception as e:
        print(f"  ❌ {name} ({interval}): Error fetching live data - {e}")
        return None, False

# ── ANALYSE ────────────────────────────────────────────────────────
def analyse(name, ticker, interval):
    df, is_live = fetch_live_data(name, ticker, interval)
    
    # Only proceed if we got live data
    if not is_live or df is None:
        return None
    
    ha = to_heikin_ashi(df.copy())
    ha["RSI"]                  = calc_rsi(ha["HA_Close"], RSI_PERIOD)
    ha["SK"], ha["SD"]         = calc_stoch(ha["HA_High"], ha["HA_Low"], ha["HA_Close"], STOCH_K, STOCH_D)
    ha["MACD"], ha["MSIG"], ha["MHIST"] = calc_macd(ha["HA_Close"])
    ha["BB_MID"],ha["BB_UP"],ha["BB_LO"] = calc_bb(ha["HA_Close"])
    ha["EMA21"] = calc_ema(ha["HA_Close"], 21)
    ha["EMA9"]  = calc_ema(ha["HA_Close"], 9)
    ha["SUPP"]  = ha["HA_Low"].rolling(20).min()
    ha["RES"]   = ha["HA_High"].rolling(20).max()
    ha["VOLAVG"]= ha["Volume"].rolling(20).mean()

    lat  = ha.iloc[-1]
    prev = ha.iloc[-2] if len(ha) > 1 else ha.iloc[-1]
    rsi  = round(float(lat["RSI"]), 2) if pd.notna(lat["RSI"]) else 50
    sk   = round(float(lat["SK"]),  2) if pd.notna(lat["SK"]) else 50
    sd   = round(float(lat["SD"]),  2) if pd.notna(lat["SD"]) else 50
    ltp  = round(float(lat["HA_Close"]), 2)
    chg  = round((float(lat["HA_Close"])-float(prev["HA_Close"]))/float(prev["HA_Close"])*100, 2) if prev["HA_Close"] != 0 else 0
    sig  = get_signal(rsi, sk)
    mh   = float(lat["MHIST"]) if pd.notna(lat["MHIST"]) else 0
    va   = float(lat["VOLAVG"]) if pd.notna(lat["VOLAVG"]) else float("nan")
    str_ = signal_strength(rsi, sk, mh, float(lat["Volume"]) if pd.notna(lat["Volume"]) else 0, va)
    sup  = round(float(lat["SUPP"]), 2) if pd.notna(lat["SUPP"]) else None
    res  = round(float(lat["RES"]),  2) if pd.notna(lat["RES"])  else None

    # Count signals in history
    signals = ha.apply(lambda r: get_signal(r["RSI"] if pd.notna(r["RSI"]) else 50, 
                                            r["SK"] if pd.notna(r["SK"]) else 50), axis=1)
    buy_candles = int((signals == "🟢 BUY").sum())
    sell_candles = int((signals == "🔴 SELL").sum())

    return {
        "Stock": name, "Ticker": ticker, "Timeframe": interval,
        "LTP (₹)": ltp, "Change (%)": chg,
        "RSI": rsi, "Stoch %K": sk, "Stoch %D": sd,
        "Signal": sig, "Strength": str_,
        "MACD Hist": round(mh, 4) if not np.isnan(mh) else 0,
        "BB Upper": round(float(lat["BB_UP"]),2) if pd.notna(lat["BB_UP"]) else None,
        "BB Lower": round(float(lat["BB_LO"]),2) if pd.notna(lat["BB_LO"]) else None,
        "Support": sup, "Resistance": res,
        "Buy Candles": buy_candles,
        "Sell Candles": sell_candles,
        "As of": ha.index[-1].strftime("%d-%b-%Y %H:%M"),
        "Data Source": "Live (Yahoo Finance)",
        "_ha": ha,
    }

# ── EXCEL HELPERS ──────────────────────────────────────────────────
def hdr_cell(cell, txt, bg=HDR_DARK, fg=WHITE, bold=True, sz=9, wrap=False):
    cell.value = txt
    cell.font  = Font(name="Arial", bold=bold, color=fg, size=sz)
    cell.fill  = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    cell.border = brd

def title_row(ws, text, ncols, row=1, bg=HDR_DARK, sz=13):
    ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
    c = ws.cell(row, 1, text)
    c.font = Font(name="Arial", bold=True, size=sz, color=WHITE)
    c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 28

def data_cell(cell, value, bg=WHITE, fg="000000", bold=False, sz=9, num_fmt=None):
    cell.value = value
    cell.font  = Font(name="Arial", size=sz, bold=bold, color=fg)
    cell.fill  = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = brd
    if num_fmt:
        cell.number_format = num_fmt

# ── GET TOP 10 STOCKS ─────────────────────────────────────────────
def get_top_10_stocks(results_by_tf):
    """Get top 10 stocks based on 1H performance (Change %)"""
    if "1h" not in results_by_tf or not results_by_tf["1h"]:
        return list(NSE_STOCKS.keys())[:10]
    
    # Sort by Change % descending and get top 10
    sorted_stocks = sorted(results_by_tf["1h"], key=lambda x: x["Change (%)"], reverse=True)
    top_10 = [r["Stock"] for r in sorted_stocks[:10]]
    print(f"\n🏆 Top 10 performing stocks (based on 1H change): {', '.join(top_10)}")
    return top_10

# ── BUILD EXCEL ────────────────────────────────────────────────────
def build_excel(results_by_tf, top_10_stocks):
    wb = Workbook()
    now_str  = datetime.now().strftime("%d-%b-%Y %H:%M")
    
    # ── Prepare combined results ──
    all_results = {}
    for tf, res_list in results_by_tf.items():
        for r in res_list:
            if r is not None:
                all_results[(r["Stock"], tf)] = r
    
    stocks = top_10_stocks
    
    # ════════════════════════════════════════════════════════════════
    # SHEET 1: Dashboard (Multi-Timeframe) - TOP 10 ONLY
    # ════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "📊 Dashboard Multi-TF"
    ws1.freeze_panes = "A5"
    ws1.sheet_view.showGridLines = False

    title_row(ws1, f"NSE TOP 10 — 1H / 4H / 1D SIGNALS  |  Updated: {now_str}", 25)
    
    ws1.merge_cells("A2:Y2")
    c2 = ws1.cell(2,1, f"🟢 BUY: RSI>{BUY_RSI} & Stoch>{BUY_STOCH}  |  🔴 SELL: RSI<{SELL_RSI} & Stoch<{SELL_STOCH}  |  Candles: Heikin Ashi  |  Showing TOP 10 stocks")
    c2.font = Font(name="Arial", size=9, italic=True, color=WHITE)
    c2.fill = PatternFill("solid", start_color=HDR_MED)
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[2].height = 18

    ws1.merge_cells("A3:Y3")
    c3 = ws1.cell(3, 1, "⚠ Educational use only. Not financial advice. Always verify with your broker.")
    c3.font = Font(name="Arial", size=8, italic=True, color="7F6000")
    c3.fill = PatternFill("solid", start_color="FFF2CC")
    c3.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[3].height = 16

    # Headers - 3 timeframes side by side
    dash_cols = ["#","Stock","Ticker"]
    tf_headers = ["1H RSI","1H SK","1H SIG","1H Str","1H Sup","1H Res",
                  "4H RSI","4H SK","4H SIG","4H Str","4H Sup","4H Res",
                  "1D RSI","1D SK","1D SIG","1D Str","1D Sup","1D Res",
                  "LTP","Chg %","Buy C","Sell C","Source"]
    
    header_row = 4
    for c, h in enumerate(dash_cols + tf_headers, 1):
        hdr_cell(ws1.cell(header_row, c), h, wrap=True, sz=8)
    ws1.row_dimensions[header_row].height = 40

    for ri, stock in enumerate(stocks, 5):
        alt = (ri % 2 == 0)
        rbg = ALT_BG if alt else WHITE
        
        # Get data for each timeframe
        tf_data = {}
        for tf in TIMEFRAMES:
            key = (stock, tf)
            tf_data[tf] = all_results.get(key, None)
        
        # Build row
        row_vals = [ri-4, stock, NSE_STOCKS[stock]]
        
        # For each timeframe
        for tf in TIMEFRAMES:
            r = tf_data.get(tf)
            if r:
                sig = r["Signal"]
                row_vals.extend([r["RSI"], r["Stoch %K"], sig, r["Strength"], 
                                 r["Support"] or "-", r["Resistance"] or "-"])
            else:
                row_vals.extend(["-", "-", "⏳ WAIT", 0, "-", "-"])
        
        # LTP, Change, Buy/Sell candles
        r1h = tf_data.get("1h")
        if r1h:
            row_vals.extend([r1h["LTP (₹)"], r1h["Change (%)"], 
                            r1h["Buy Candles"], r1h["Sell Candles"], r1h["Data Source"]])
        else:
            row_vals.extend(["-", 0, 0, 0, "-"])
        
        # Write row
        col_idx = 1
        for val in row_vals:
            cell = ws1.cell(ri, col_idx, val)
            # Check if this is a signal column (3rd, 9th, 15th positions in tf blocks)
            sig_cols = [5, 11, 17]  # 1H SIG, 4H SIG, 1D SIG
            str_cols = [6, 12, 18]  # 1H Str, 4H Str, 1D Str
            
            if col_idx in sig_cols:
                sbg, sfg = sig_style(val)
                data_cell(cell, val, sbg, sfg, bold=True)
            elif col_idx in str_cols:
                if isinstance(val, (int, float)):
                    s = val
                else:
                    try:
                        s = float(val) if val != "-" else 0
                    except:
                        s = 0
                str_bg = STR_HI_BG if s >= 60 else (STR_MED_BG if s >= 30 else STR_LO_BG)
                data_cell(cell, val, str_bg, "000000", bold=True)
            else:
                data_cell(cell, val, rbg)
            
            if col_idx == 21 and isinstance(val, float):  # Chg %
                cell.number_format = '+0.00%;-0.00%'
            col_idx += 1
        
        ws1.row_dimensions[ri].height = 22

    col_widths = [4, 13, 14, 7, 7, 10, 6, 9, 9, 7, 7, 10, 6, 9, 9, 7, 7, 10, 6, 10, 8, 7, 7, 14]
    for i, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ════════════════════════════════════════════════════════════════
    # SHEETS 2-4: History for each timeframe - TOP 10 ONLY
    # ════════════════════════════════════════════════════════════════
    tf_names = {"1h": "1H", "4h": "4H", "1d": "1D"}
    
    for tf in TIMEFRAMES:
        sheet_name = f"📈 {tf_names[tf]} History"
        ws = wb.create_sheet(sheet_name)
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A4"
        title_row(ws, f"{tf_names[tf]} HEIKIN ASHI — TOP 10 STOCKS  |  Updated: {now_str}", 11)
        
        ws.merge_cells("A2:K2")
        c2b = ws.cell(2,1,"Signal: BUY (RSI>40 & Stoch>20) | SELL (RSI<70 & Stoch<80)")
        c2b.font = Font(name="Arial", size=9, italic=True, color=WHITE)
        c2b.fill = PatternFill("solid", start_color=HDR_MED)
        c2b.alignment = Alignment(horizontal="center")
        ws.row_dimensions[2].height = 18

        h2cols = ["Stock","Date/Time","HA Open","HA Close","HA High","HA Low",
                  "RSI","Stoch %K","Stoch %D","MACD Hist","Signal"]
        for c, h in enumerate(h2cols, 1):
            hdr_cell(ws.cell(3, c), h, wrap=True)
        ws.row_dimensions[3].height = 28

        row = 4
        for stock in stocks:
            r = all_results.get((stock, tf))
            if r is None:
                continue
            ha = r["_ha"].tail(20)
            for _, rec in ha.iterrows():
                sig = get_signal(rec["RSI"] if pd.notna(rec["RSI"]) else 50,
                                 rec["SK"] if pd.notna(rec["SK"]) else 50)
                sbg, sfg = sig_style(sig)
                mh_val = round(float(rec["MHIST"]),4) if pd.notna(rec.get("MHIST")) else "-"
                vals = [
                    stock,
                    rec.name.strftime("%d-%b %H:%M"),
                    round(float(rec["HA_Open"]),2),
                    round(float(rec["HA_Close"]),2),
                    round(float(rec["HA_High"]),2),
                    round(float(rec["HA_Low"]),2),
                    round(float(rec["RSI"]),2) if pd.notna(rec["RSI"]) else "-",
                    round(float(rec["SK"]),2)  if pd.notna(rec["SK"])  else "-",
                    round(float(rec["SD"]),2)  if pd.notna(rec["SD"])  else "-",
                    mh_val, sig
                ]
                alt2 = row % 2 == 0
                rbg  = ALT_BG if alt2 else WHITE
                for ci, v in enumerate(vals, 1):
                    cell = ws.cell(row, ci, v)
                    if ci == 11:
                        data_cell(cell, v, sbg, sfg, bold=True, sz=9)
                    else:
                        data_cell(cell, v, rbg, "000000", sz=9)
                ws.row_dimensions[row].height = 18
                row += 1

        for i, w in enumerate([13,14,10,10,10,10,7,10,10,12,14], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ════════════════════════════════════════════════════════════════
    # SHEET 5: Charts - TOP 10 ONLY
    # ════════════════════════════════════════════════════════════════
    ws5 = wb.create_sheet("📊 Charts")
    ws5.sheet_view.showGridLines = False
    title_row(ws5, f"TOP 10 — RSI & STRENGTH  |  Updated: {now_str}", 7)
    ws5.row_dimensions[1].height = 28

    # Write chart data - RSI comparison
    chart_row = 3
    ws5.cell(chart_row, 1, "Stock").font = Font(bold=True, size=9, name="Arial")
    for ci, tf in enumerate(TIMEFRAMES, 2):
        ws5.cell(chart_row, ci, f"{tf_names[tf]} RSI").font = Font(bold=True, size=9, name="Arial")
        ws5.cell(chart_row, ci+3, f"{tf_names[tf]} Strength").font = Font(bold=True, size=9, name="Arial")
    
    for c in range(1, 8):
        cell = ws5.cell(chart_row, c)
        cell.fill = PatternFill("solid", start_color=HDR_DARK)
        cell.font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    for ri, stock in enumerate(stocks, chart_row+1):
        ws5.cell(ri, 1, stock)
        for ci, tf in enumerate(TIMEFRAMES, 2):
            r = all_results.get((stock, tf))
            if r:
                ws5.cell(ri, ci, r["RSI"])
                ws5.cell(ri, ci+3, r["Strength"])
            else:
                ws5.cell(ri, ci, "-")
                ws5.cell(ri, ci+3, "-")

    n = len(stocks)
    data_end = chart_row + n

    # RSI Chart
    chart_rsi = BarChart()
    chart_rsi.type = "col"
    chart_rsi.title = "RSI (14) Comparison — TOP 10"
    chart_rsi.y_axis.title = "RSI Value"
    chart_rsi.x_axis.title = "Stock"
    chart_rsi.style = 10
    chart_rsi.width = 28
    chart_rsi.height = 16
    
    data_rsi = Reference(ws5, min_col=2, min_row=chart_row, max_col=4, max_row=data_end)
    cats = Reference(ws5, min_col=1, min_row=chart_row+1, max_row=data_end)
    chart_rsi.add_data(data_rsi, titles_from_data=True)
    chart_rsi.set_categories(cats)
    ws5.add_chart(chart_rsi, "I3")

    # Strength Chart
    chart_str = BarChart()
    chart_str.type = "col"
    chart_str.title = "Signal Strength — TOP 10"
    chart_str.y_axis.title = "Strength Score"
    chart_str.x_axis.title = "Stock"
    chart_str.style = 10
    chart_str.width = 28
    chart_str.height = 16
    
    data_str = Reference(ws5, min_col=5, min_row=chart_row, max_col=7, max_row=data_end)
    chart_str.add_data(data_str, titles_from_data=True)
    chart_str.set_categories(cats)
    ws5.add_chart(chart_str, "I33")

    ws5.column_dimensions["A"].width = 14
    for i in range(2, 8):
        ws5.column_dimensions[get_column_letter(i)].width = 12

    # ════════════════════════════════════════════════════════════════
    # SHEET 6: Leaderboard - TOP 10 ONLY
    # ════════════════════════════════════════════════════════════════
    ws6 = wb.create_sheet("🏆 Leaderboard")
    ws6.sheet_view.showGridLines = False
    title_row(ws6, f"TOP 10 — SIGNAL STRENGTH LEADERBOARD  |  Updated: {now_str}", 9)
    
    ws6.merge_cells("A2:I2")
    c4b = ws6.cell(2,1,"Scores: 60–100 = Strong ✅  |  30–59 = Moderate ⚠  |  0–29 = Weak ❌")
    c4b.font = Font(name="Arial", size=9, italic=True, color=WHITE)
    c4b.fill = PatternFill("solid", start_color=HDR_MED)
    c4b.alignment = Alignment(horizontal="center")
    ws6.row_dimensions[2].height = 18

    lb_cols = ["Rank","Stock","Ticker","TF","Signal","Score","RSI","Stoch %K","LTP (₹)"]
    for c, h in enumerate(lb_cols, 1):
        hdr_cell(ws6.cell(3, c), h, wrap=True)
    ws6.row_dimensions[3].height = 28

    # Flatten results for leaderboard
    flat_results = []
    for tf, res_list in results_by_tf.items():
        for r in res_list:
            if r is not None and r["Stock"] in top_10_stocks:
                flat_results.append(r)
    flat_results.sort(key=lambda x: x["Strength"], reverse=True)

    for ri, r in enumerate(flat_results, 4):
        sbg, sfg = sig_style(r["Signal"])
        s = r["Strength"]
        str_bg = STR_HI_BG if s >= 60 else (STR_MED_BG if s >= 30 else STR_LO_BG)
        alt3 = ri % 2 == 0
        rbg = ALT_BG if alt3 else WHITE

        vals = [ri-3, r["Stock"], r["Ticker"], r["Timeframe"], r["Signal"], 
                r["Strength"], r["RSI"], r["Stoch %K"], r["LTP (₹)"]]
        
        medal = {4:"🥇", 5:"🥈", 6:"🥉"}.get(ri, "")
        if medal: vals[0] = medal

        for ci, v in enumerate(vals, 1):
            cell = ws6.cell(ri, ci, v)
            if ci == 5:
                data_cell(cell, v, sbg, sfg, bold=True)
            elif ci == 6:
                data_cell(cell, v, str_bg, "000000", bold=True)
            else:
                data_cell(cell, v, rbg, "000000", bold=(ci <= 2))
        ws6.row_dimensions[ri].height = 22

    for i, w in enumerate([6, 13, 14, 6, 14, 10, 8, 10, 11], 1):
        ws6.column_dimensions[get_column_letter(i)].width = w

    # ════════════════════════════════════════════════════════════════
    # SHEET 7: Strategy Guide (Updated)
    # ════════════════════════════════════════════════════════════════
    ws7 = wb.create_sheet("📋 Guide")
    ws7.sheet_view.showGridLines = False
    title_row(ws7, "NSE TOP 10 — HEIKIN ASHI STRATEGY GUIDE", 2)

    guide = [
        ("WHAT IS HEIKIN ASHI?", None, HDR_MED, WHITE, 10, True),
        ("HA Candles", "Uses averaged OHLC values — smooths noise and makes trends clearer than regular candlesticks.", "EEF2FF","000000",10,False),
        ("HA Close", "(Open+High+Low+Close)÷4 — used for all indicator calculations.", WHITE,"000000",10,False),
        ("Solid Body","All-green body (no lower wick) = strong uptrend. All-red body (no upper wick) = strong downtrend.", "EEF2FF","000000",10,False),
        ("EMA 9 & 21","Short EMA (9) crossing above long EMA (21) = bullish momentum.", WHITE,"000000",10,False),
        ("","","FFFFFF","000000",9,False),
        ("SIGNAL RULES (UPDATED)", None, HDR_MED, WHITE, 10, True),
        (f"🟢 BUY",f"RSI > {BUY_RSI}  AND  Stoch %K > {BUY_STOCH}  → Bullish momentum. Look for HA color change to green.", BUY_BG, BUY_FG, 10, True),
        (f"🔴 SELL",f"RSI < {SELL_RSI}  AND  Stoch %K < {SELL_STOCH}  → Bearish momentum. Look for HA color change to red.", SELL_BG, SELL_FG, 10, True),
        (f"🟡 HOLD","Neither condition met. Wait for signal confirmation.", HOLD_BG, HOLD_FG, 10, True),
        ("","","FFFFFF","000000",9,False),
        ("TOP 10 SELECTION", None, HDR_MED, WHITE, 10, True),
        ("Method","Stocks are sorted by 1H Change % (performance). Only top 10 performers are shown.", "EEF2FF","000000",10,False),
        ("Purpose","Focus on momentum stocks that are actively moving in the market.", WHITE,"000000",10,False),
        ("","","FFFFFF","000000",9,False),
        ("MULTI-TIMEFRAME APPROACH", None, HDR_MED, WHITE, 10, True),
        ("1H","Short-term momentum. Best for entry/exit timing. More frequent signals.", "EEF2FF","000000",10,False),
        ("4H","Medium-term trend. Balances noise and signal quality. Use for swing trades.", WHITE,"000000",10,False),
        ("1D","Long-term trend. Major trend direction. Use for position sizing and bias.", "EEF2FF","000000",10,False),
        ("","","FFFFFF","000000",9,False),
        ("STRENGTH SCORE", None, HDR_MED, WHITE, 10, True),
        ("60–100","STRONG — RSI & Stoch both deep in zone, MACD confirms. Act on this signal.", STR_HI_BG,"276221",10,False),
        ("30–59", "MODERATE — Partial confirmation. Exercise caution. Wait for next candle.", STR_MED_BG,"9C5700",10,False),
        ("0–29",  "WEAK — Avoid trading on this signal alone. Wait for better setup.", STR_LO_BG,"9C0006",10,False),
        ("","","FFFFFF","000000",9,False),
        ("INDICATOR DETAILS", None, HDR_MED, WHITE, 10, True),
        ("RSI (14)","Relative Strength Index. Calculated on HA Close.", "EEF2FF","000000",10,False),
        ("Stochastic","K=14, D=3. Uses HA High/Low/Close.", WHITE,"000000",10,False),
        ("MACD","Fast=12, Slow=26, Signal=9. Histogram above 0 = bullish momentum.", "EEF2FF","000000",10,False),
        ("","","FFFFFF","000000",9,False),
        ("⚠ RISK WARNINGS", None, HDR_MED, WHITE, 10, True),
        ("Timeframes","Higher timeframes (1D) have stronger signals. Lower timeframes (1H) have more noise.", "FFF2CC","7F6000",10,False),
        ("Confirmation","Best signals occur when all 3 timeframes align (same direction).", "FFF2CC","7F6000",10,False),
        ("Not Advice","Educational only. Do your own research. Consult a SEBI-registered advisor.", "FFF2CC","7F6000",10,False),
        ("Stop Loss","Always set a stop-loss. Use HA swing low/high for entries.", "FFF2CC","7F6000",10,False),
        ("Live Data","ONLY live data from Yahoo Finance (.NS suffix). No simulation used.", "FFF2CC","7F6000",10,False),
    ]

    for i, (key, val, bg, fg, sz, bold) in enumerate(guide, 2):
        c1 = ws7.cell(i, 1, key)
        c1.font = Font(name="Arial", size=sz, bold=bold, color=fg)
        c1.fill = PatternFill("solid", start_color=bg)
        c1.alignment = Alignment(vertical="center", wrap_text=True)
        c1.border = brd
        ws7.row_dimensions[i].height = 22
        if val is not None:
            c2 = ws7.cell(i, 2, val)
            c2.font = Font(name="Arial", size=sz, bold=bold, color=fg)
            c2.fill = PatternFill("solid", start_color=bg)
            c2.alignment = Alignment(vertical="center", wrap_text=True)
            c2.border = brd
        else:
            ws7.merge_cells(f"A{i}:B{i}")
            ws7.cell(i,1).alignment = Alignment(horizontal="center", vertical="center")

    ws7.column_dimensions["A"].width = 24
    ws7.column_dimensions["B"].width = 75

    # ── Save ───────────────────────────────────────────────────────
    script_dir = os.path.dirname(os.path.abspath(sys.argv[0]))
    out = os.path.join(script_dir, OUTPUT_FILE)
    wb.save(out)
    print(f"\n✅ Excel saved → {out}")
    return out

# ── MAIN ───────────────────────────────────────────────────────────
if __name__ == "__main__":
    print(f"🔍 NSE Analyzer v6  |  Timeframes: 1H, 4H, 1D  |  Candles: Heikin Ashi")
    print(f"📊 Signal Rules: BUY (RSI>{BUY_RSI} & Stoch>{BUY_STOCH}) | SELL (RSI<{SELL_RSI} & Stoch<{SELL_STOCH})")
    print(f"🏆 Showing TOP 10 stocks by 1H performance\n")
    
    results_by_tf = {}
    
    for tf in TIMEFRAMES:
        print(f"\n📊 Analysing {tf} timeframe...")
        tf_results = []
        success_count = 0
        for name, ticker in NSE_STOCKS.items():
            try:
                r = analyse(name, ticker, tf)
                if r is not None:
                    tf_results.append(r)
                    success_count += 1
                    bar = "█"*int(r["Strength"]//10) + "░"*(10-int(r["Strength"]//10))
                    print(f"  {name:12s}  ₹{r['LTP (₹)']:>9.2f}  RSI:{r['RSI']:>5.1f}  "
                          f"SK:{r['Stoch %K']:>5.1f}  [{bar}] {r['Strength']:>3d}  {r['Signal']}")
                else:
                    print(f"  {name:12s}  ❌ No live data available")
            except Exception as e:
                print(f"  {name}: ERROR — {e}")
        results_by_tf[tf] = tf_results
        print(f"  ✅ {success_count}/{len(NSE_STOCKS)} stocks analysed successfully for {tf}")

    # Get top 10 stocks based on 1H performance
    top_10_stocks = get_top_10_stocks(results_by_tf)

    print(f"\n📁 Building Excel (7 sheets, native charts) for TOP 10 stocks …")
    build_excel(results_by_tf, top_10_stocks)
